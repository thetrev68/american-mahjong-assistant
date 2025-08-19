#!/usr/bin/env python3
"""
Convert existing JSON complete hands to Excel template
Creates a clean Excel file with 687 working hands for manual editing
"""

import json
import pandas as pd
import openpyxl
import openpyxl.styles
import openpyxl.utils
from typing import List, Dict, Any

def load_existing_hands(json_file: str) -> List[Dict]:
    """Load the existing complete hands JSON"""
    with open(json_file, 'r') as f:
        data = json.load(f)
    return data['complete_hands']

def json_hands_to_excel_rows(hands: List[Dict]) -> List[Dict]:
    """Convert JSON hands to Excel row format"""
    excel_rows = []
    
    for hand in hands:
        # Skip failed generations
        if not hand.get('generation_success', True):
            continue
            
        # Skip hands with wrong tile count
        if hand['exact_tiles']['total_tiles'] != 14:
            continue
        
        # Start building the Excel row
        row = {
            'hand_id': hand['hand_id'],
            'pattern_key': hand['pattern_info']['pattern_key'],
            'display_pattern': hand['pattern_info']['display_pattern'],
            'description': hand['pattern_info']['description'],
            'points': hand['pattern_info']['points'],
            'difficulty': hand['pattern_info']['difficulty'],
            'concealed': hand['pattern_info']['concealed']
        }
        
        # Extract tile information
        required_tiles = hand['exact_tiles']['required_tiles']
        
        # Get joker rules - create lookup by tile position
        joker_lookup = {}
        for group in hand['joker_rules']['groups']:
            jokers_allowed = group['jokers_allowed']
            group_tiles = group['tiles']
            
            # Map each tile in this group to its joker allowance
            for tile in group_tiles:
                joker_lookup[tile] = jokers_allowed
        
        # Fill in 14 tile positions
        for i in range(1, 15):  # 1-14
            if i-1 < len(required_tiles):
                tile_id = required_tiles[i-1]
                joker_allowed = joker_lookup.get(tile_id, True)  # Default to True if not specified
                
                row[f'tile_{i}_id'] = tile_id
                row[f'tile_{i}_joker'] = joker_allowed
            else:
                # Empty tile position
                row[f'tile_{i}_id'] = ''
                row[f'tile_{i}_joker'] = False
        
        excel_rows.append(row)
    
    return excel_rows

def create_excel_template(json_file: str, output_file: str):
    """Create Excel template from JSON hands"""
    
    print("Loading existing complete hands...")
    hands = load_existing_hands(json_file)
    print(f"Loaded {len(hands)} hands from JSON")
    
    print("Converting to Excel format...")
    excel_rows = json_hands_to_excel_rows(hands)
    print(f"Converted {len(excel_rows)} successful hands")
    
    # Create DataFrame
    df = pd.DataFrame(excel_rows)
    
    # Define column order
    core_columns = ['hand_id', 'pattern_key', 'display_pattern', 'description', 'points', 'difficulty', 'concealed']
    
    # Tile columns in order
    tile_columns = []
    for i in range(1, 15):
        tile_columns.extend([f'tile_{i}_id', f'tile_{i}_joker'])
    
    # Reorder columns
    column_order = core_columns + tile_columns
    df = df.reindex(columns=column_order)
    
    # Create Excel writer with formatting
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Complete_Hands', index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Complete_Hands']
        
        # Format headers
        header_font = openpyxl.styles.Font(bold=True, color='FFFFFF')
        header_fill = openpyxl.styles.PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        for col in range(1, len(column_order) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
                    
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add data validation for joker columns
        from openpyxl.worksheet.datavalidation import DataValidation
        
        joker_validation = DataValidation(
            type="list",
            formula1='"TRUE,FALSE"',
            allow_blank=False
        )
        
        # Apply to all joker columns for all rows
        for i in range(1, 15):
            joker_col = column_order.index(f'tile_{i}_joker') + 1  # +1 for Excel indexing
            joker_col_letter = openpyxl.utils.get_column_letter(joker_col)
            
            joker_validation.add(f'{joker_col_letter}2:{joker_col_letter}{len(excel_rows) + 1}')
        
        worksheet.add_data_validation(joker_validation)
    
    print(f"Excel template created: {output_file}")
    print(f"Contains {len(excel_rows)} hands ready for manual editing")
    
    # Print summary statistics
    pattern_counts = {}
    for row in excel_rows:
        key = row['pattern_key']
        pattern_counts[key] = pattern_counts.get(key, 0) + 1
    
    print(f"\nPattern distribution:")
    for pattern, count in sorted(pattern_counts.items()):
        print(f"  {pattern}: {count} hands")

def main():
    """Main execution"""
    try:
        import openpyxl
    except ImportError:
        print("Error: openpyxl is required. Install with: pip install openpyxl")
        return
    
    json_file = "frontend/public/intelligence/nmjl-patterns/nmjl-complete-hands-2025.json"
    output_file = "nmjl-complete-hands-template.xlsx"
    
    try:
        create_excel_template(json_file, output_file)
        print(f"\nSuccess! Excel template ready at: {output_file}")
        print("\nYou can now:")
        print("1. Edit the tile positions and joker rules as needed")
        print("2. Add more hand variations manually")
        print("3. Use the conversion script to generate final JSON")
        
    except FileNotFoundError as e:
        print(f"Error: Could not find file: {e}")
    except Exception as e:
        print(f"Error creating Excel template: {e}")

if __name__ == "__main__":
    main()